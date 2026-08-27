"""Builds the client-facing test workbook from the automated run."""

import csv
import datetime as dt
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

RESULTS_TSV = sys.argv[1]
OUT = sys.argv[2]

NAVY = "0B2C3D"
TEAL = "0E9F76"
GREY = "F1EDE6"
RED = "DC2626"
AMBER = "B45309"

HEAD_FILL = PatternFill("solid", fgColor=NAVY)
HEAD_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=16, color=NAVY)
WRAP = Alignment(vertical="top", wrap_text=True)
TOP = Alignment(vertical="top")
THIN = Side(style="thin", color="D8D2C8")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = Workbook()


def style_header(ws, row=1):
    for cell in ws[row]:
        if cell.value is not None:
            cell.fill = HEAD_FILL
            cell.font = HEAD_FONT
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w


def add_status_validation(ws, col_letter, first, last):
    """Pass/Fail/Blocked/Not tested dropdown, so results come back consistent."""
    dv = DataValidation(
        type="list",
        formula1='"Pass,Fail,Blocked,Not tested"',
        allow_blank=True,
        showDropDown=False,
    )
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{first}:{col_letter}{last}")


# ── Sheet 1: Read me ─────────────────────────────────────────────────────────
ws = wb.active
ws.title = "Read me"
widths(ws, {"A": 26, "B": 78})

ws["A1"] = "RESET — Test Pack"
ws["A1"].font = TITLE_FONT
ws["A2"] = f"Generated {dt.date.today().isoformat()}"
ws["A2"].font = Font(italic=True, color="6B7280")

rows = [
    ("", ""),
    ("WHERE TO TEST", ""),
    ("Customer website", "https://resetmen.in"),
    ("Admin panel", "https://admin.resetmen.in"),
    ("API (health check)", "https://api.resetmen.in/api/v1/health/ready"),
    ("Android app", "Installed on the test device. Not yet on the Play Store."),
    ("", ""),
    ("ADMIN LOGIN", ""),
    ("Email", "admin@resetmen.in"),
    ("Password", "Reset@123"),
    ("Role", "OWNER — full access"),
    (
        "IMPORTANT",
        "This password is weak and is for testing only. Change it before the store "
        "goes live to real customers.",
    ),
    ("", ""),
    ("CUSTOMER LOGIN", ""),
    (
        "How to sign in",
        "Sign in with Google in the app or on the website. Any Google account works — "
        "no account needs to be created for you.",
    ),
    (
        "Phone number required",
        "A booking cannot be made until a phone number is saved on the profile. This is "
        "deliberate: the store rings the customer to take payment.",
    ),
    ("", ""),
    ("HOW PAYMENT WORKS TODAY", ""),
    (
        "No online payment",
        "There is no payment gateway connected. Every booking is created unpaid and the "
        "money is taken at the counter.",
    ),
    (
        "Collecting money",
        "Admin panel > Payments due lists every booking for a day, what is outstanding, "
        "and the customer's phone number. Ring them, take payment, press Mark paid.",
    ),
    ("", ""),
    ("HOW TO USE THIS PACK", ""),
    (
        "1",
        "Work through the 'Customer app tests' and 'Admin panel tests' sheets in order.",
    ),
    ("2", "Put Pass or Fail in the Result column. Add anything odd in Notes."),
    (
        "3",
        "The 'Automated results' sheet is a record of checks already run against the live "
        "system — no action needed, it is there for reference.",
    ),
    ("", ""),
    ("KNOWN GAPS", ""),
    (
        "No photographs",
        "No service photos have been uploaded, so the app shows coloured icons instead. "
        "Uploading photos in the admin panel replaces them, with no app update needed.",
    ),
    (
        "Instant Glow is empty",
        "The category exists with no services in it, so it is hidden. It needs its "
        "services adding before it will appear.",
    ),
    (
        "SMS and push",
        "No SMS provider is connected and push notifications are not configured, so "
        "these are recorded but not delivered.",
    ),
]

r = 3
for label, value in rows:
    ws.cell(row=r, column=1, value=label)
    ws.cell(row=r, column=2, value=value).alignment = WRAP
    if label.isupper() and label:
        ws.cell(row=r, column=1).font = Font(bold=True, color=TEAL, size=11)
    elif label in ("IMPORTANT",):
        ws.cell(row=r, column=1).font = Font(bold=True, color=RED)
        ws.cell(row=r, column=2).font = Font(color=RED)
    else:
        ws.cell(row=r, column=1).font = Font(bold=True)
    ws.cell(row=r, column=1).alignment = TOP
    r += 1


# ── Manual test case sheets ──────────────────────────────────────────────────
def manual_sheet(title, cases):
    s = wb.create_sheet(title)
    headers = ["ID", "Area", "Test case", "Steps", "Expected result", "Result", "Notes"]
    s.append(headers)
    for c in cases:
        s.append(list(c) + ["", ""])
    widths(s, {"A": 8, "B": 16, "C": 34, "D": 52, "E": 46, "F": 12, "G": 30})
    for row in s.iter_rows(min_row=2, max_row=s.max_row, max_col=7):
        for cell in row:
            cell.alignment = WRAP
            cell.border = BOX
    style_header(s)
    add_status_validation(s, "F", 2, s.max_row)
    s.auto_filter.ref = f"A1:G{s.max_row}"
    return s


app_cases = [
    ("A01", "Install", "App opens", "Tap the RESET icon on the phone.",
     "The app opens on the Book tab without crashing."),
    ("A02", "Catalogue", "Services are listed", "Look at the Book tab.",
     "Stress Relief and Full Body Relax appear, each with services, prices and durations."),
    ("A03", "Catalogue", "Prices are correct",
     "Compare the prices shown against the store's own price list.",
     "Head 49, Head+Neck+Shoulder 99, +Back 149, Basic 199, Premium 299 (rupees)."),
    ("A04", "Catalogue", "Pull to refresh", "Pull down on the Book tab.",
     "The list refreshes without error."),
    ("A05", "Search", "Search filters the list", "Tap the search box and type 'head'.",
     "Only services with 'head' in the name or description remain."),
    ("A06", "Search", "Search can be cleared", "Type a search, then press the X.",
     "The full list returns."),
    ("A07", "Search", "No match is handled", "Search for 'zzzz'.",
     "A 'no match' message appears rather than a blank screen."),
    ("A08", "Categories", "Category filter works", "Tap the Stress Relief circle.",
     "Only Stress Relief services are shown. Tapping it again shows everything."),
    ("A09", "Service", "Service page opens", "Tap any service.",
     "The service page opens with its picture, name, price and duration."),
    ("A10", "Service", "Add-ons can be chosen", "On a service with add-ons, select one.",
     "The price and duration update to include it."),
    ("A11", "Sign in", "Google sign-in works", "Tap You, then sign in with Google.",
     "The Google account chooser appears and sign-in completes."),
    ("A12", "Profile", "Phone can be saved", "In You, add a phone number and save.",
     "The number is saved and still there after closing and reopening the app."),
    ("A13", "Booking", "Booking is blocked without a phone",
     "Sign in with an account that has no phone number, then try to book.",
     "The app asks for a phone number and does not create the booking."),
    ("A14", "Slots", "Times are listed", "Choose a service, then a date.",
     "Available times appear in a grid."),
    ("A15", "Slots", "Closed day shows nothing", "Pick a Monday.",
     "No times are offered — the store is shut on Mondays."),
    ("A16", "Slots", "Only 7 days ahead", "Try to pick a date more than a week away.",
     "Dates beyond a week are not offered."),
    ("A17", "Booking", "A booking can be made", "Pick a time and press Book.",
     "The booking is created and the confirmation screen appears."),
    ("A18", "Booking", "The button says Book, not Pay", "Look at the button on checkout.",
     "It reads 'Book' and the screen says payment is taken at the counter."),
    ("A19", "Booking", "No money is requested", "Complete a booking.",
     "No payment screen appears at any point."),
    ("A20", "Booking", "QR code is shown", "Open the booking after making it.",
     "A QR code is displayed for check-in at the store."),
    ("A21", "Visits", "Booking appears in Visits", "Open the Visits tab.",
     "The new booking is listed with its date, time and status."),
    ("A22", "Visits", "Booking can be cancelled", "Open a booking and cancel it.",
     "The booking shows as Cancelled."),
    ("A23", "Rewards", "Rewards tab opens", "Open the Rewards tab.",
     "Streak progress and scratch cards load without error."),
    ("A24", "Navigation", "All four tabs work", "Tap Book, Visits, Rewards, You in turn.",
     "Each opens without crashing."),
    ("A25", "Navigation", "Back returns, does not exit",
     "Open a service, then press the phone's back button.",
     "You return to the list. The app does not close."),
    ("A26", "Navigation", "Back from another tab returns to Book",
     "Open Rewards, then press back.",
     "You land on the Book tab rather than leaving the app."),
    ("A27", "Navigation", "Tabs remember their place",
     "Open a service, switch to Rewards, switch back to Book.",
     "You are still on the service you had open."),
    ("A28", "Offline", "No connection is handled", "Turn on aeroplane mode, open the app.",
     "A clear message appears. The app does not crash."),
    ("A29", "Offline", "Booking QR still works offline",
     "With a booking made, turn on aeroplane mode and open Visits.",
     "The booking and its QR code are still visible."),
    ("A30", "Display", "Large text is handled",
     "Set the phone's font size to the largest, reopen the app.",
     "Text is readable and nothing important is cut off."),
]

admin_cases = [
    ("D01", "Login", "Admin can sign in",
     "Go to https://admin.resetmen.in and sign in with the credentials on the Read me sheet.",
     "The dashboard loads."),
    ("D02", "Login", "Wrong password is refused", "Try signing in with a wrong password.",
     "Sign-in is refused with a clear message."),
    ("D03", "Login", "Signed-out access is blocked",
     "Sign out, then go straight to https://admin.resetmen.in/timeline.",
     "You are sent back to the sign-in page."),
    ("D04", "Dashboard", "Today's numbers load", "Open the dashboard.",
     "Today's bookings and revenue figures appear."),
    ("D05", "Timeline", "The day is shown by station", "Open Timeline.",
     "Each station is shown with its bookings laid out through the day."),
    ("D06", "Timeline", "Another date can be chosen", "Change the date.",
     "That day's bookings load."),
    ("D07", "Payments due", "Unpaid bookings are listed", "Open Payments due.",
     "Every unpaid booking for the day appears with the amount owed."),
    ("D08", "Payments due", "Outstanding total is right",
     "Compare the Outstanding figure with the rows listed.",
     "The total matches the sum of the unpaid rows."),
    ("D09", "Payments due", "Customer phone is shown", "Look at the Phone column.",
     "The customer's number is shown and can be tapped to call."),
    ("D10", "Payments due", "Money can be recorded",
     "Choose a method (Cash/UPI/Card), then press Mark paid on a booking.",
     "The row moves to Paid and the Collected total increases."),
    ("D11", "Payments due", "Marking twice is safe", "Press Mark paid twice on the same booking.",
     "It says it was already paid. The takings do not double."),
    ("D12", "Payments due", "Paid bookings can be reviewed", "Press 'Show all'.",
     "Paid bookings appear alongside unpaid ones, marked Paid."),
    ("D13", "Walk-in", "A walk-in can be entered",
     "From the timeline, add a walk-in for someone at the counter.",
     "The booking is created and occupies a station."),
    ("D14", "Check in", "A booking can be checked in",
     "Open Check in and enter or scan a booking code.",
     "The booking is marked as checked in."),
    ("D15", "Catalog", "A service can be edited",
     "Open Catalog, change a service's price, save.",
     "The new price is shown, and appears in the app after a refresh."),
    ("D16", "Catalog", "A service photo can be uploaded",
     "Open Catalog, upload an image for a service.",
     "The photo appears in the app in place of the coloured icon."),
    ("D17", "Catalog", "A service can be hidden",
     "Set a service to inactive.",
     "It disappears from the app."),
    ("D18", "Capacity", "Stations are listed", "Open Capacity.",
     "The three stations are listed with the services each can take."),
    ("D19", "Capacity", "Opening hours can be changed",
     "Change an opening or closing time and save.",
     "Available times in the app change to match."),
    ("D20", "Capacity", "A closed day can be set",
     "Mark a day as closed.",
     "No times are offered for that day in the app."),
    ("D21", "Customers", "Customers are listed", "Open Customers.",
     "Customers appear with their contact details and booking history."),
    ("D22", "Staff", "A staff member can be added",
     "Open Staff and add a member with the Staff role.",
     "They can sign in and see only the counter screens."),
    ("D23", "Reports", "Revenue report runs",
     "Open Reports and run the revenue report for a date range.",
     "Figures appear and can be exported."),
    ("D24", "Rewards", "Reward rules are listed", "Open Rewards.",
     "Streak rules and scratch card campaigns are shown."),
    ("D25", "Audit", "Actions are recorded", "Open Audit log after marking a booking paid.",
     "The action appears with who did it and when."),
    ("D26", "Roles", "Staff cannot reach owner screens",
     "Sign in as a Staff user and try to open Staff or Audit log.",
     "Access is refused."),
]

manual_sheet("Customer app tests", app_cases)
manual_sheet("Admin panel tests", admin_cases)


# ── Automated results ────────────────────────────────────────────────────────
auto = wb.create_sheet("Automated results")
auto.append(["ID", "Area", "Check", "Expected", "Actual", "Result", "Notes"])

summary = None
with open(RESULTS_TSV, newline="", encoding="utf-8") as fh:
    for row in csv.reader(fh, delimiter="\t"):
        if not row:
            continue
        if row[0] == "SUMMARY":
            summary = row
            continue
        auto.append(row)

widths(auto, {"A": 8, "B": 14, "C": 44, "D": 16, "E": 16, "F": 10, "G": 28})
for row in auto.iter_rows(min_row=2, max_row=auto.max_row, max_col=7):
    for cell in row:
        cell.alignment = TOP
        cell.border = BOX
    result = row[5].value
    if result == "PASS":
        row[5].font = Font(color=TEAL, bold=True)
    elif result == "FAIL":
        row[5].font = Font(color=RED, bold=True)
        for cell in row:
            cell.fill = PatternFill("solid", fgColor="FDECEC")
style_header(auto)
auto.auto_filter.ref = f"A1:G{auto.max_row}"

if summary is not None:
    last = auto.max_row + 2
    auto.cell(row=last, column=1, value="TOTAL").font = Font(bold=True)
    auto.cell(row=last, column=3, value=f"{summary[3]} checks run against the live system")
    auto.cell(row=last, column=4, value=f"{summary[1]} passed").font = Font(color=TEAL, bold=True)
    auto.cell(row=last, column=5, value=f"{summary[2]} failed").font = Font(
        color=RED if summary[2] != "0" else TEAL, bold=True
    )
    auto.cell(row=last + 1, column=3, value=f"Run on {dt.date.today().isoformat()}").font = Font(
        italic=True, color="6B7280"
    )

wb.save(OUT)
print(f"written: {OUT}")
print(f"sheets: {', '.join(wb.sheetnames)}")
